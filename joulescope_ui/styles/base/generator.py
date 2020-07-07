# Copyright 2020 Jetperch LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from openpyxl import load_workbook
import re
import os

MYPATH = os.path.dirname(os.path.abspath(__file__))
WBPATH = os.path.join(MYPATH, 'colors.xlsm')


STYLES = ['disabled', 'enabled', 'hover', 'pressed']


SVG_CHECKABLE = [
    # must provide *_checked.svg and *_unchecked.svg
    # generate disabled, enabled, hover, pressed
    'radio',
    'checkbox'
]

SVG_CLICKABLE = [
    'arrow_down',
    'arrow_left',
    'arrow_right',
    'arrow_up',
    'close',
    'undock',
    'zoom_all',
    'zoom_in',
    'zoom_out',
]

SVG_SIMPLE = [
    'hmovetoolbar',
    'hsepartoolbar',
    'vmovetoolbar',
    'vsepartoolbars',
    'sizegrip',
    'branch_closed',
    'branch_end',
    'branch_end_open',
    'branch_more',
    'branch_open',
    'branch_vline',
    'transparent',
]

SVG_COPY = [
    'play',
    'record',
    'record_statistics'
]


def load_colors():
    worbook = load_workbook(filename=WBPATH)
    sheet = worbook['colors']
    colors = {}
    for row in sheet.iter_rows(min_row=2, max_row=1000, min_col=1, max_col=3, values_only=True):
        if not row[0]:
            continue
        color_name = row[0]
        colors[color_name] = row[1]
    return colors


def generate_svg(svg_names, colors, target_path):
    filenames = []
    for svg_name in svg_names:
        with open(os.path.join(MYPATH, f'{svg_name}.svg'), encoding='utf-8') as f:
            svg = f.read()
        for style in colors.keys():
            out_name = f'{svg_name}_{style}.svg'
            svg_out = svg.replace('#000000', colors[style])
            fname = os.path.join(target_path, out_name)
            with open(fname, 'w', encoding='utf-8') as f:
                f.write(svg_out)
            filenames.append(out_name)
    return filenames


def copy_svg(svg_names, target_path):
    filenames = []
    for svg_name in svg_names:
        fname = f'{svg_name}.svg'
        with open(os.path.join(MYPATH, fname), encoding='utf-8') as f:
            svg = f.read()
        with open(os.path.join(target_path, fname), 'w', encoding='utf-8') as f:
            f.write(svg)
        filenames.append(fname)
    return filenames


def generate_css(colors, target_path):
    fname = 'style.qss'
    r = re.compile(r'{%\s*([a-zA-Z0-9_]+)\s*%}')
    def replace(matchobj):
        return colors[matchobj.group(1)]
    with open(os.path.join(MYPATH, fname), encoding='utf-8') as f:
        css = f.read()
    css = r.sub(replace, css)
    with open(os.path.join(target_path, fname), 'w', encoding='utf-8') as f:
        f.write(css)
    return [fname]


def generate_qrc(filenames, target_path):
    parts = ['<RCC>\n  <qresource prefix="style">']
    for fname in filenames:
        parts.append(f'    <file>{fname}</file>')
    parts.append('  </qresource>\n</RCC>\n')
    with open(os.path.join(target_path, 'style.qrc'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(parts))


def run():
    filenames = []
    target = 'demo'
    target_path = os.path.join(os.path.dirname(MYPATH), target)
    if not os.path.isdir(target_path):
        os.makedirs(target_path, exist_ok=True)
    colors = load_colors()
    # print(colors)

    unchecked_colors = {}
    checked_colors = {}
    clickable_colors = {}
    for style in STYLES:
        unchecked_colors[style] = colors[f'unchecked_{style}']
        checked_colors[style] = colors[f'checked_{style}']
        clickable_colors[style] = colors[f'clickable_{style}']
    simple_colors = {
        'disabled': colors['foreground_disabled'],
        'enabled': colors['foreground'],
    }
    filenames += generate_svg([f'{n}_unchecked' for n in SVG_CHECKABLE], unchecked_colors, target_path)
    filenames += generate_svg([f'{n}_checked' for n in SVG_CHECKABLE], checked_colors, target_path)
    filenames += generate_svg(SVG_CLICKABLE, clickable_colors, target_path)
    filenames += generate_svg(SVG_SIMPLE, simple_colors, target_path)
    filenames += copy_svg(SVG_COPY, target_path)
    filenames += generate_css(colors, target_path)
    generate_qrc(filenames, target_path)


if __name__ == '__main__':
    run()
