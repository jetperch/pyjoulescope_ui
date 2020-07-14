# Copyright 2020 Jetperch LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from openpyxl import load_workbook
import json
import copy
import re
import os

MYPATH = os.path.dirname(os.path.abspath(__file__))
WBPATH = os.path.join(MYPATH, 'colors.xlsm')
INDEX_PATH = os.path.join(MYPATH, 'index.json')


STYLES = ['disabled', 'enabled', 'hover', 'pressed']


SVG_CHECKABLE = [
    # must provide *_checked.svg and *_unchecked.svg
    # generate disabled, enabled, hover, pressed
    'radio',
    'checkbox'
]

SVG_CLICKABLE = [
    'arrow_down',
    'arrow_left',
    'arrow_right',
    'arrow_up',
    'close',
    'undock',
    'zoom_all',
    'zoom_in',
    'zoom_out',
]

SVG_SIMPLE = [
    'hmovetoolbar',
    'hsepartoolbar',
    'vmovetoolbar',
    'vsepartoolbars',
    'sizegrip',
    'branch_closed',
    'branch_end',
    'branch_end_open',
    'branch_more',
    'branch_open',
    'branch_vline',
    'transparent',
]

SVG_COPY = [
    'play',
    'record',
    'record_statistics'
]


class Generator:

    def __init__(self, path, css_path=None, subtheme=None):
        self.filenames = []
        with open(INDEX_PATH, 'r', encoding='utf-8') as f:
            self.index = json.load(f)
        subthemes = self.subthemes
        if subtheme is not None:
            if subtheme not in subthemes:
                raise ValueError(f'subtheme {subtheme} not in {subthemes}')
        else:
            subtheme = subthemes[0]
        self.subtheme = subtheme

        path = os.path.normpath(path)
        dname = f"{self.index['name']}_{subtheme}"
        self._path = os.path.join(os.path.normpath(path), dname)
        if css_path is None:
            css_path = self._path.replace('\\', '/')
        if len(css_path) and css_path[-1] != '/':
            css_path += '/'
        self._css_path = css_path

    def make_path(self):
        if not os.path.isdir(self._path):
            os.makedirs(self._path, exist_ok=True)

    @property
    def subthemes(self):
        return list(self.index['colors'].keys())

    def generate_svg(self, svg_names, colors):
        filenames = []
        for svg_name in svg_names:
            with open(os.path.join(MYPATH, f'{svg_name}.svg'), encoding='utf-8') as f:
                svg = f.read()
            for style in colors.keys():
                out_name = f'{svg_name}_{style}.svg'
                svg_out = svg.replace('#000000', colors[style])
                fname = os.path.join(self._path, out_name)
                with open(fname, 'w', encoding='utf-8') as f:
                    f.write(svg_out)
                filenames.append(out_name)
        self.filenames.extend(filenames)
        return filenames

    def copy_svg(self, svg_names):
        filenames = []
        for svg_name in svg_names:
            fname = f'{svg_name}.svg'
            with open(os.path.join(MYPATH, fname), encoding='utf-8') as f:
                svg = f.read()
            with open(os.path.join(self._path, fname), 'w', encoding='utf-8') as f:
                f.write(svg)
            filenames.append(fname)
        self.filenames.extend(filenames)
        return filenames

    def generate_css(self):
        fname = 'style.qss'
        r = re.compile(r'{%\s*([a-zA-Z0-9_]+)\s*%}')
        colors = self.index['colors'][self.subtheme]
        def replace(matchobj):
            varname = matchobj.group(1)
            if varname == 'path':
                return self._css_path
            return colors[varname]
        with open(os.path.join(MYPATH, fname), encoding='utf-8') as f:
            css = f.read()
        css = r.sub(replace, css)
        with open(os.path.join(self._path, fname), 'w', encoding='utf-8') as f:
            f.write(css)
        self.filenames.append(fname)
        return [fname]

    def generate_index(self):
        index = copy.deepcopy(self.index)
        index['colors'] = index['colors'][self.subtheme]
        index['name'] = f"{index['name']}_{self.subtheme}"
        index['description'] += f" - {self.subtheme}"
        with open(os.path.join(self._path, 'index.json'), 'w', encoding='utf-8') as f:
            json.dump(index, f, indent=2)

    def generate_qrc(self):
        parts = ['<RCC>\n  <qresource prefix="style">']
        for fname in self.filenames:
            parts.append(f'    <file>{fname}</file>')
        parts.append('  </qresource>\n</RCC>\n')
        with open(os.path.join(self._path, 'style.qrc'), 'w', encoding='utf-8') as f:
            f.write('\n'.join(parts))

    def run(self):
        colors = self.index['colors'][self.subtheme]
        unchecked_colors = {}
        checked_colors = {}
        clickable_colors = {}
        for style in STYLES:
            unchecked_colors[style] = colors[f'unchecked_{style}']
            checked_colors[style] = colors[f'checked_{style}']
            clickable_colors[style] = colors[f'clickable_{style}']
        simple_colors = {
            'disabled': colors['disabled_foreground'],
            'enabled': colors['base_foreground'],
        }
        self.make_path()
        self.generate_svg([f'{n}_unchecked' for n in SVG_CHECKABLE], unchecked_colors)
        self.generate_svg([f'{n}_checked' for n in SVG_CHECKABLE], checked_colors)
        self.generate_svg(SVG_CLICKABLE, clickable_colors)
        self.generate_svg(SVG_SIMPLE, simple_colors)
        self.copy_svg(SVG_COPY)
        self.generate_css()
        self.generate_index()
        # self.generate_qrc()


def index_update():
    dark = {}
    light = {}
    with open(INDEX_PATH, 'r', encoding='utf-8') as f:
        index = json.load(f)

    workbook = load_workbook(filename=WBPATH)
    sheet = workbook['colors']

    for row in sheet.iter_rows(min_row=2, max_row=1000, min_col=1, max_col=5, values_only=True):
        if not row[0]:
            continue
        color_name = row[0]
        dark[color_name] = row[1]
        light[color_name] = row[3]
    index['colors'] = {
        'dark': dark,
        'light': light,
    }

    with open(INDEX_PATH, 'w', encoding='utf-8') as f:
        json.dump(index, f, indent=2)


def run():
    path = 'C:/Users/Matth/AppData/Local/joulescope/themes/'
    #target = 'demo'
    #target_path = os.path.join(os.path.dirname(MYPATH), target)
    index_update()
    generator = Generator(path=path)
    generator.run()


if __name__ == '__main__':
    run()
